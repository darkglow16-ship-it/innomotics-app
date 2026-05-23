from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from io import BytesIO
from copy import copy
import base64, os

app = Flask(__name__)
CORS(app)

BASE_DIR  = os.path.dirname(__file__)
PLANTILLA = os.path.join(BASE_DIR, 'plantilla_sem.xlsx')

PLANILLAS = [
    'CODIGOS_CORREAS_PUERTO.xlsx',
    'LUBRICANTES_PATACHE.xlsx',
    'Matriz_de_riesgo_310FL00xxxx.xlsx',
    'Planilla_2026_Inventario_Puerto_Patache_CMDIC.csv',
]

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/planillas')
def listar_planillas():
    return jsonify({'planillas': PLANILLAS})

@app.route('/planilla/<nombre>')
def servir_planilla(nombre):
    if nombre not in PLANILLAS:
        return jsonify({'error': 'Planilla no encontrada'}), 404
    ruta = os.path.join(BASE_DIR, nombre)
    if not os.path.exists(ruta):
        return jsonify({'error': 'Archivo no encontrado'}), 404
    with open(ruta, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    return jsonify({'nombre': nombre, 'data': b64})

@app.route('/generar', methods=['POST'])
def generar():
    try:
        datos  = request.get_json()
        semana = datos['semana']
        año    = datos['año']
        filas  = datos['filas']

        wb = load_workbook(PLANTILLA)
        ws = wb.active

        ws['O2'] = semana
        ws['H4'] = f'COMPRAS SEM {semana} INNOMOTICS PUERTO'

        def copy_styles(row_num):
            s = {}
            for cell in ws[row_num]:
                s[cell.column] = {
                    'font':      copy(cell.font),
                    'border':    copy(cell.border),
                    'fill':      copy(cell.fill),
                    'alignment': copy(cell.alignment),
                }
            return s

        s8 = copy_styles(8)
        s9 = copy_styles(9)

        for r in range(8, 83):
            for c in range(2, 16):
                ws.cell(row=r, column=c).value = None

        for i, f in enumerate(filas):
            r   = 8 + i
            sty = s8 if i == 0 else s9
            ws.row_dimensions[r].height = 19.35

            def wr(col, val, fmt=None, aln=None):
                cell = ws.cell(row=r, column=col)
                cell.value = val
                st = sty.get(col, {})
                for attr in ('font','border','fill','alignment'):
                    if st.get(attr): setattr(cell, attr, copy(st[attr]))
                if fmt: cell.number_format = fmt
                if aln: cell.alignment = aln

            wr(2, i+1)
            if f.get('fechaSol'): wr(3, datetime.strptime(f['fechaSol'],'%Y-%m-%d'), 'DD/MM/YYYY')
            if f.get('fechaReq'): wr(4, datetime.strptime(f['fechaReq'],'%Y-%m-%d'), 'DD/MM/YYYY')
            wr(5, datetime.strptime(f['fechaEjec'],'%Y-%m-%d'), 'DD/MM/YYYY')
            wr(6, f['ot'])
            wr(7, f['equipo'])
            wr(8, f['grupo'])
            wr(9, f['cs'])
            wr(10, int(f['qty']) if str(f.get('qty',1)).isdigit() else 1)
            wr(11, f['desc'], aln=Alignment(horizontal='left', vertical='center', wrap_text=True))
            wr(12, '')
            wr(13, f.get('stock', ''))
            wr(14, f.get('comentario', ''))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()
        return jsonify({'ok': True, 'b64': b64})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
